# gui/views/reports/export_utils.py
"""
Utilities for exporting reports to various formats.

This module provides functions for exporting reports to PDF, Excel,
and other formats, as well as printing functionality.
"""

import logging
import os
import tempfile
import tkinter as tk
from tkinter import filedialog, messagebox
from typing import Any, Dict, List, Optional, Tuple, Union
from datetime import datetime
import csv

# Try to import openpyxl for Excel export, but provide fallback to CSV
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logging.warning("openpyxl not available, Excel export will use CSV format")

logger = logging.getLogger(__name__)


class ReportExporter:
    """Handles exporting reports to various formats."""

    @staticmethod
    def export_to_pdf(report_title: str, report_data: List[Dict],
                      columns: List[Dict], filename: Optional[str] = None,
                      include_summary: bool = True,
                      summary_data: Optional[Dict[str, Any]] = None) -> bool:
        """
        Export report data to PDF.

        Args:
            report_title: The title of the report
            report_data: List of dictionaries containing report data
            columns: List of column definitions (name, key, width)
            filename: Optional filename to save to (if None, user will be prompted)
            include_summary: Whether to include summary data
            summary_data: Dictionary of summary metrics to include

        Returns:
            True if export was successful, False otherwise
        """
        try:
            # This is a placeholder for actual PDF generation
            # In a real implementation, we would use ReportLab

            if not filename:
                # Ask user for filename if not provided
                default_filename = get_default_report_filename(report_title, "pdf")
                filename = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                    title="Save PDF Report",
                    initialfile=default_filename
                )

                if not filename:  # User cancelled
                    return False

            # Log that we would create a PDF here
            logger.info(f"Exporting report '{report_title}' to PDF: {filename}")
            logger.info(f"Report contains {len(report_data)} rows and {len(columns)} columns")

            # In a real implementation, this would generate the actual PDF
            # For now, we'll just create a dummy file
            with open(filename, 'w') as f:
                f.write(f"PDF Report: {report_title}\n")
                f.write(f"Generated: {datetime.now()}\n\n")

                # Include summary if requested
                if include_summary and summary_data:
                    f.write("Summary:\n")
                    for key, value in summary_data.items():
                        f.write(f"{key}: {value}\n")
                    f.write("\n")

                # Write headers
                header_line = ""
                for col in columns:
                    header_line += f"{col['name']}\t"
                f.write(f"{header_line}\n")

                # Write data
                for row in report_data:
                    data_line = ""
                    for col in columns:
                        data_line += f"{row.get(col['key'], '')}\t"
                    f.write(f"{data_line}\n")

                f.write("\n(This is a placeholder file for demonstration purposes)")

            messagebox.showinfo("Export Successful",
                                f"Report exported to {os.path.basename(filename)}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to PDF: {str(e)}", exc_info=True)
            messagebox.showerror("Export Failed",
                                 f"Failed to export report: {str(e)}")
            return False

    @staticmethod
    def export_to_excel(report_title: str, report_data: List[Dict],
                        columns: List[Dict], filename: Optional[str] = None,
                        include_summary: bool = True,
                        summary_data: Optional[Dict[str, Any]] = None) -> bool:
        """
        Export report data to Excel format.

        Args:
            report_title: The title of the report
            report_data: List of dictionaries containing report data
            columns: List of column definitions (name, key, width)
            filename: Optional filename to save to (if None, user will be prompted)
            include_summary: Whether to include summary data
            summary_data: Dictionary of summary metrics to include

        Returns:
            True if export was successful, False otherwise
        """
        try:
            if not filename:
                # Ask user for filename if not provided
                default_filename = get_default_report_filename(report_title, "xlsx" if HAS_OPENPYXL else "csv")
                filename = filedialog.asksaveasfilename(
                    defaultextension=".xlsx" if HAS_OPENPYXL else ".csv",
                    filetypes=[
                        ("Excel files", "*.xlsx") if HAS_OPENPYXL else ("CSV files", "*.csv"),
                        ("All files", "*.*")
                    ],
                    title=f"Save {'Excel' if HAS_OPENPYXL else 'CSV'} Report",
                    initialfile=default_filename
                )

                if not filename:  # User cancelled
                    return False

            # Log export information
            logger.info(f"Exporting report '{report_title}' to {'Excel' if HAS_OPENPYXL else 'CSV'}: {filename}")

            if HAS_OPENPYXL:
                # Export using openpyxl for proper Excel format
                return ReportExporter._export_to_excel_openpyxl(
                    report_title, report_data, columns, filename, include_summary, summary_data
                )
            else:
                # Fall back to CSV export
                return ReportExporter._export_to_csv(
                    report_title, report_data, columns, filename, include_summary, summary_data
                )

        except Exception as e:
            logger.error(f"Error exporting to Excel/CSV: {str(e)}", exc_info=True)
            messagebox.showerror("Export Failed",
                                 f"Failed to export report: {str(e)}")
            return False

    @staticmethod
    def _export_to_excel_openpyxl(report_title: str, report_data: List[Dict],
                                  columns: List[Dict], filename: str,
                                  include_summary: bool = True,
                                  summary_data: Optional[Dict[str, Any]] = None) -> bool:
        """
        Export to Excel using openpyxl library.

        Args:
            report_title: The title of the report
            report_data: List of dictionaries containing report data
            columns: List of column definitions (name, key, width)
            filename: Filename to save to
            include_summary: Whether to include summary data
            summary_data: Dictionary of summary metrics to include

        Returns:
            True if export was successful, False otherwise
        """
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Add report title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = report_title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Add generation timestamp
        ws.merge_cells('A2:F2')
        date_cell = ws['A2']
        date_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.alignment = Alignment(horizontal='center')

        # Add summary if requested
        current_row = 3
        if include_summary and summary_data:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            summary_title = ws[f'A{current_row}']
            summary_title.value = "Summary"
            summary_title.font = Font(bold=True)
            current_row += 1

            for key, value in summary_data.items():
                ws[f'A{current_row}'] = key
                ws[f'B{current_row}'] = value
                current_row += 1

            # Add a spacer row
            current_row += 1

        # Add headers
        header_row = current_row
        header_style = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = col['name']
            cell.font = header_style
            cell.fill = header_fill

            # Set column width (approximation from pixels to Excel units)
            excel_width = min(max(col.get('width', 100) / 7, 10), 60)  # Limit between 10-60 units
            ws.column_dimensions[get_column_letter(col_idx)].width = excel_width

        # Add data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row_idx, row_data in enumerate(report_data, header_row + 1):
            for col_idx, col in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(col['key'], '')

                # Format special types
                if isinstance(value, datetime):
                    cell.value = value
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                elif isinstance(value, (int, float)) and col.get('key', '').lower().endswith(
                        ('price', 'cost', 'value')):
                    cell.value = value
                    cell.number_format = '$#,##0.00'
                else:
                    cell.value = value

                cell.border = thin_border

        # Save the workbook
        wb.save(filename)

        messagebox.showinfo("Export Successful",
                            f"Report exported to {os.path.basename(filename)}")
        return True

    @staticmethod
    def _export_to_csv(report_title: str, report_data: List[Dict],
                       columns: List[Dict], filename: str,
                       include_summary: bool = True,
                       summary_data: Optional[Dict[str, Any]] = None) -> bool:
        """
        Export to CSV format (fallback when openpyxl is not available).

        Args:
            report_title: The title of the report
            report_data: List of dictionaries containing report data
            columns: List of column definitions (name, key, width)
            filename: Filename to save to
            include_summary: Whether to include summary data
            summary_data: Dictionary of summary metrics to include

        Returns:
            True if export was successful, False otherwise
        """
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)

            # Write report title and timestamp
            writer.writerow([report_title])
            writer.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])  # Empty row

            # Write summary if requested
            if include_summary and summary_data:
                writer.writerow(["Summary"])
                for key, value in summary_data.items():
                    writer.writerow([key, value])
                writer.writerow([])  # Empty row

            # Write headers
            writer.writerow([col['name'] for col in columns])

            # Write data
            for row in report_data:
                csv_row = []
                for col in columns:
                    value = row.get(col['key'], '')
                    # Format dates
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    csv_row.append(value)
                writer.writerow(csv_row)

        messagebox.showinfo("Export Successful",
                            f"Report exported to {os.path.basename(filename)}")
        return True

    @staticmethod
    def print_report(report_title: str, report_data: List[Dict],
                     columns: List[Dict],
                     include_summary: bool = True,
                     summary_data: Optional[Dict[str, Any]] = None) -> bool:
        """
        Print the report.

        Args:
            report_title: The title of the report
            report_data: List of dictionaries containing report data
            columns: List of column definitions (name, key, width)
            include_summary: Whether to include summary data
            summary_data: Dictionary of summary metrics to include

        Returns:
            True if print was successful, False otherwise
        """
        try:
            # This is a placeholder for actual printing functionality
            # In a real implementation, we would use a library for printing

            # Create a temporary file for printing
            with tempfile.NamedTemporaryFile(delete=False, suffix='.txt', mode='w') as f:
                temp_filename = f.name
                f.write(f"PRINT: {report_title}\n")
                f.write(f"Generated: {datetime.now()}\n\n")

                # Include summary if requested
                if include_summary and summary_data:
                    f.write("Summary:\n")
                    for key, value in summary_data.items():
                        f.write(f"{key}: {value}\n")
                    f.write("\n")

                # Write headers
                header_line = ""
                for col in columns:
                    header_line += f"{col['name']}\t"
                f.write(f"{header_line}\n")

                # Write data
                for row in report_data:
                    data_line = ""
                    for col in columns:
                        value = row.get(col['key'], '')
                        # Format dates
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                        data_line += f"{value}\t"
                    f.write(f"{data_line}\n")

            logger.info(f"Print job prepared for report '{report_title}'")
            logger.info(f"Temp file created at {temp_filename}")

            # In a real implementation, this would send the file to the printer
            # For now, we'll just notify the user
            messagebox.showinfo("Print Job Prepared",
                                "The report is ready to print. In a real implementation, "
                                "this would send the report to the printer.")

            # Clean up the temporary file
            os.unlink(temp_filename)
            return True

        except Exception as e:
            logger.error(f"Error preparing print job: {str(e)}", exc_info=True)
            messagebox.showerror("Print Failed",
                                 f"Failed to prepare print job: {str(e)}")
            return False


def get_default_report_filename(report_title: str, extension: str = "pdf") -> str:
    """
    Generate a default filename for a report based on title and current date.

    Args:
        report_title: The title of the report
        extension: The file extension to use

    Returns:
        A formatted filename
    """
    # Sanitize the report title for use in a filename
    sanitized_title = "".join(c if c.isalnum() or c in [' ', '_'] else '_' for c in report_title)
    sanitized_title = sanitized_title.replace(' ', '_')

    # Get current date and time
    now = datetime.now()
    date_str = now.strftime("%Y%m%d_%H%M%S")

    return f"{sanitized_title}_{date_str}.{extension}"